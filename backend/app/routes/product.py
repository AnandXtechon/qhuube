from fastapi import APIRouter, HTTPException, Depends
from app.core.security import verify_access_token
from app.schemas.product_schemas import ProductSchema, ProductCreateSchema, ProductListResponse
from app.models.product_model import get_all_products, create_product, update_product, delete_product
from fastapi import UploadFile, File
from openpyxl import load_workbook
from io import BytesIO

router = APIRouter()

@router.post("/create/product", response_model=ProductSchema)
async def create_new_product(product: ProductCreateSchema, admin=Depends(verify_access_token)):

    product_data = await create_product(
        product_type=product.product_type,
        country=product.country,
        vat_rate=product.vat_rate,
        vat_category=product.vat_category,
        shipping_vat_rate=product.shipping_vat_rate,
    )

    product_data["_id"] = str(product_data["_id"])
    
    return ProductSchema(**product_data)



@router.get("/products", response_model=ProductListResponse)
async def get_products(admin=Depends(verify_access_token)):
    products = await get_all_products()
    return {
        "success": True,
        "products": products
    }


@router.put("/update/product/{product_id}", response_model=ProductSchema)
async def update_existing_product(product_id: str, updated: ProductCreateSchema, admin=Depends(verify_access_token)):
    try:
        updated_product = await update_product(
            product_id=product_id,
            product_type=updated.product_type,
            country=updated.country,
            vat_rate=updated.vat_rate,
            vat_category=updated.vat_category,
            shipping_vat_rate=updated.shipping_vat_rate
        )
        
        if "_id" in updated_product:
            updated_product["_id"] = str(updated_product["_id"])

        return ProductSchema(**updated_product)
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))
    

@router.delete("/delete/product/{product_id}")
async def delete_existing_product(product_id: str, admin=Depends(verify_access_token)):
    try:
        return await delete_product(product_id)
    except Exception as e:
        raise HTTPException(status_code=404, detail=str(e))
    

# Import products from Excel file
# @router.post("/imports/products")
# async def import_products(file: UploadFile = File(...), admin=Depends(verify_access_token)):
#     # 1. Check file type
#     if not file.filename.endswith((".xlsx", ".xls")):
#         raise HTTPException(status_code=400, detail="File must be .xlsx or .xls format")

#     try:
#         # 2. Read Excel content
#         content = await file.read()
#         workbook = load_workbook(BytesIO(content))
#         sheet = workbook.active

#         # 3. Validate header row
#         headers = [cell.value for cell in sheet[1]]
#         expected_headers = ["product_type", "country", "vat_rate", "vat_category", "shipping_vat_rate"]
#         if headers != expected_headers:
#             raise HTTPException(status_code=400, detail=f"Expected headers: {expected_headers}, but got: {headers}")

#         # 4. Process rows
#         success_count = 0
#         errors = []

#         for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
#             try:
#                 product_data = ProductCreateSchema(
#                     product_type=row[0].value,
#                     country=row[1].value,
#                     vat_rate=float(row[2].value),
#                     vat_category=row[3].value,
#                     shipping_vat_rate=float(row[4].value)
#                 )

#                 await create_product(
#                     product_type=product_data.product_type,
#                     country=product_data.country,
#                     vat_rate=product_data.vat_rate,
#                     vat_category=product_data.vat_category,
#                     shipping_vat_rate=product_data.shipping_vat_rate
#                 )
#                 success_count += 1

#             except Exception as e:
#                 errors.append(f"Row {idx}: {str(e)}")

#         return {
#             "success": True,
#             "imported": success_count,
#             "errors": errors
#         }

#     except Exception as e:
#         raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")
   

# Import products from JSON
@router.post("/imports/products/json")
async def import_products_json(data: dict, admin=Depends(verify_access_token)):
    products = data.get("products", [])
    success_count = 0
    errors = []

    for idx, row in enumerate(products, start=1):
        try:
            product = ProductCreateSchema(**row)
            await create_product(
                product_type=product.product_type,
                country=product.country,
                vat_rate=product.vat_rate,
                vat_category=product.vat_category,
                shipping_vat_rate=product.shipping_vat_rate,
            )
            success_count += 1
        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    return {
        "imported": success_count,
        "errors": errors
    }
