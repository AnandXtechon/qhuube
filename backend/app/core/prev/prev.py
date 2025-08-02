from io import BytesIO
import io
from typing import List
import pandas as pd
import numpy as np
from fastapi import BackgroundTasks, Form, UploadFile, HTTPException, APIRouter, File
from fastapi.responses import JSONResponse, StreamingResponse
from app.models.header_model import get_all_headers
from app.models.product_model import get_all_products
from app.core.helper import safe_float, safe_round, dataframe_to_json_safe, get_user_friendly_dtype, TYPE_MAP
from app.core.currency_conversion import get_ecb_fx_rates, get_fx_rate_for_date
from app.core.send_mail import send_manual_vat_email
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook, load_workbook
import re

router = APIRouter()


# Read uploaded file and return headers + DataFrame
async def extract_file_headers(file: UploadFile) -> tuple[list[str], pd.DataFrame]:
    try:
        content = await file.read()
        file_data = BytesIO(content)
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file_data)
        elif file.filename.endswith('.txt'):
            df = pd.read_csv(file_data, delimiter='\t')
        else:
            df = pd.read_excel(file_data)

        headers = [str(col).strip().lower() for col in df.columns]

        return headers, df
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error reading file: {str(e)}"
        )


async def validate_file_data(file_headers: list[str], df: pd.DataFrame) -> dict:
    try:
        all_headers = await get_all_headers()

        alias_to_value = {}
        required_headers = []
        header_labels = {}
        expected_types = {}

        for header in all_headers:
            value = header['value']
            required_headers.append(value)
            header_labels[value] = header['label']
            for alias in header['aliases']:
                alias_to_value[alias.strip().lower()] = value
            
             # Get the raw type from database
            raw_type = header.get('type', 'string')
            print(f"Header: {value}, Raw type from DB: {raw_type}")
            
            # Map the type using TYPE_MAP
            mapped_type = TYPE_MAP.get(raw_type.lower(), 'string')
            expected_types[value] = mapped_type

        rename_map = {}

        # for col in range(len(file_headers)):
        #     normalized = file_headers[col].strip().lower()
        #     mapped_value = alias_to_value.get(normalized)
        #     if mapped_value:
        #         file_headers[col] = mapped_value

        for col in df.columns:
            normalized = col.strip().lower()
            mapped_value = alias_to_value.get(normalized)
            if mapped_value:
                rename_map[col] = mapped_value

        df.rename(columns=rename_map, inplace=True)
        print("DateFrame Columns", df.columns.to_list())

        # Create detailed missing headers info
        missing_headers_detailed = []
        for field in required_headers:
            if field not in df.columns:
                missing_headers_detailed.append({
                    'header_value': field,
                    'header_label': header_labels.get(field, field),
                    'expected_name': header_labels.get(field, field),
                    'description': f"Required column '{header_labels.get(field, field)}' is missing from the file"
                })

        data_issues = []
        for header_value in df.columns:
            if header_value not in df.columns:
                continue

            col_dtype = get_user_friendly_dtype(df[header_value].dtype)

            try:
                null_mask = df[header_value].isnull()
                empty_mask = df[header_value].astype(str).str.strip().isin(['', 'nan', 'None', '(empty)', '(null)'])
                combined_mask = null_mask | empty_mask

                null_count = int(null_mask.sum())
                empty_count = int(empty_mask.sum())
                total_empty = int(combined_mask.sum())

                if total_empty > 0:
                    missing_rows = df.index[combined_mask].tolist()
                    missing_rows_display = [str(row + 2) for row in missing_rows]

                    issue_description = f"Column '{header_labels.get(header_value, header_value)}' has {total_empty} missing values"

                    if len(missing_rows) > 10:
                        issue_description += f" (showing first 10 rows: {','.join(missing_rows_display[:10])}...)"
                    else:
                        issue_description += f" in rows: {', '.join(missing_rows_display)}"

                    data_issues.append({
                        'header_value': header_value,
                        'header_label': header_labels.get(header_value, header_value),
                        'original_column': header_value,
                        'issue_type': 'MISSING_DATA',
                        'issue_description': issue_description,
                        'column_name': header_labels.get(header_value, header_value),
                        'data_type': col_dtype,
                        'null_count': null_count,
                        'empty_count': empty_count,
                        'total_missing': total_empty,
                        'percentage': round((total_empty / len(df)) * 100, 2),
                        'missing_rows': missing_rows_display,
                        'has_more_rows': len(missing_rows) > 10
                    })
            
            except Exception as col_error:
                print(f"Error processing missing data for column {header_value}: {str(col_error)}")
            
               # ------------------ Enhanced Data Type Validation ------------------
            try:
                invalid_type_rows = []
                expected_type = expected_types.get(header_value, "string")
                
                print(f"Validating column '{header_value}' against expected type: {expected_type}")
                
                # Get sample of data for debugging
                sample_data = df[header_value].dropna().head(5).tolist()
                print(f"Sample data: {sample_data}")
                
                for idx, val in df[header_value].items():
                    # Skip null/empty values as they're handled separately
                    if pd.isnull(val) or (isinstance(val, str) and val.strip() == ''):
                        continue
                    
                    is_valid = True
                    error_msg = ""
                    
                    try:
                        if expected_type == 'integer':
                            # Check if it's a boolean first
                            if isinstance(val, bool):
                                is_valid = False
                                error_msg = "Boolean value found where integer expected"
                            else:
                                # Try to convert to int
                                if isinstance(val, str):
                                    # Remove whitespace and check if it's a valid integer string
                                    clean_val = val.strip()
                                    if not clean_val.replace('-', '').replace('+', '').isdigit():
                                        is_valid = False
                                        error_msg = f"Non-integer string: '{val}'"
                                    else:
                                        int(clean_val)
                                else:
                                    # For numeric types, check if it's a whole number
                                    float_val = float(val)
                                    if float_val != int(float_val):
                                        is_valid = False
                                        error_msg = f"Decimal value where integer expected: {val}"
                                    else:
                                        int(float_val)
                        
                        elif expected_type == 'float':
                            try:
                                float(val)
                            except (ValueError, TypeError):
                                is_valid = False
                                error_msg = f"Cannot convert to float: '{val}'"
                        
                        elif expected_type == 'date':
                            # Try multiple date formats
                            date_formats = ["%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]
                            date_parsed = False
                            
                            for fmt in date_formats:
                                try:
                                    pd.to_datetime(str(val), format=fmt, errors='raise')
                                    date_parsed = True
                                    break
                                except:
                                    continue
                            
                            if not date_parsed:
                                try:
                                    pd.to_datetime(str(val), errors='raise')
                                    date_parsed = True
                                except:
                                    is_valid = False
                                    error_msg = f"Invalid date format: '{val}'"
                                    
                        
                        elif expected_type == 'text_only':
                            # New validation: text that shouldn't contain only numbers
                            str_val = str(val).strip()
                            # Check if the value is purely numeric (including decimals)
                            if re.match(r'^-?\d+\.?\d*$', str_val):
                                is_valid = False
                                error_msg = f"Numeric value found where text expected: '{val}'"
                        
                        elif expected_type == 'string':
                            # Enhanced string validation
                            str_val = str(val).strip()
                            
                            # Check for business logic rules based on column name
                            if header_value == 'product_type':
                                # Special validation for product_type - shouldn't be purely numeric
                                if re.match(r'^-?\d+\.?\d*$', str_val):
                                    is_valid = False
                                    error_msg = f"Product type should not be a number: '{val}'"
                                # Could also check against known product types
                                elif len(str_val) < 2:
                                    is_valid = False
                                    error_msg = f"Product type too short: '{val}'"
                            
                            elif header_value == 'country':
                                # Country names shouldn't be numbers
                                if re.match(r'^-?\d+\.?\d*$', str_val):
                                    is_valid = False
                                    error_msg = f"Country should not be a number: '{val}'"
                            
                            elif header_value == 'currency':
                                # Currency codes should be 3 letters
                                if not re.match(r'^[A-Z]{3}$', str_val.upper()):
                                    is_valid = False
                                    error_msg = f"Invalid currency code format: '{val}' (should be 3 letters like EUR, USD)"
                            
                            # General string validation - just ensure it can be converted to string
                            try:
                                str(val)
                            except:
                                is_valid = False
                                error_msg = f"Cannot convert to string: '{val}'"
                    
                    except Exception as validation_error:
                        is_valid = False
                        error_msg = f"Validation error: {str(validation_error)}"
                    
                    if not is_valid:
                        print(f"Type validation failed for value '{val}' at row {idx + 2}: {error_msg}")
                        invalid_type_rows.append(idx + 2)  # +2 for 1-indexed + header row
                
                if invalid_type_rows:
                    invalid_rows_display = invalid_type_rows[:10]
                    issue_description = f"Column '{header_labels.get(header_value, header_value)}' has invalid {expected_type} values in rows: {', '.join(map(str, invalid_rows_display))}"
                    if len(invalid_type_rows) > 10:
                        issue_description += "..."
                    
                    data_issues.append({
                        'header_value': header_value,
                        'header_label': header_labels.get(header_value, header_value),
                        'original_column': header_value,  # Now using standardized name
                        'issue_type': 'INVALID_TYPE',
                        'issue_description': issue_description,
                        'column_name': header_labels.get(header_value, header_value),
                        'expected_type': expected_type,
                        'invalid_rows': invalid_rows_display,
                        'invalid_count': len(invalid_type_rows),
                        'total_rows': len(df),
                        'percentage': round((len(invalid_type_rows) / len(df)) * 100, 2),
                        'has_more_rows': len(invalid_type_rows) > 10
                    })
            
            except Exception as type_error:
                print(f"Error during type validation for column {header_value}: {str(type_error)}")
        return {
            'missing_headers': [field for field in required_headers if field not in df.columns],
            'missing_headers_detailed': missing_headers_detailed,
            'matched_columns': {v: v for v in df.columns},  # Now both key and value are standardized
            'header_labels': header_labels,
            'data_issues': data_issues,
            'total_rows': len(df),
        }
    except Exception as e:
        print(f"Validation error: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Validation error: {str(e)}"
        )


async def enrich_dataframe_with_vat(df: pd.DataFrame) -> pd.DataFrame:
    try:
        # 1. Get VAT products from database
        vat_products = await get_all_products()
        print(f"Retrieved {len(vat_products)} VAT products from database")
        
        # 2. Build a lookup dictionary for (product_type, country) -> VAT info
        vat_lookup = {}
        for prod in vat_products:
            try:
                product_type = str(prod.get('product_type', '')).strip().lower()
                country = str(prod.get('country', '')).strip().lower()
                if product_type and country:
                    key = (product_type, country)
                    vat_lookup[key] = prod
                    print(f"Added VAT lookup: {key} -> VAT: {prod.get('vat_rate', 0)}%, Shipping VAT: {prod.get('shipping_vat_rate', 0)}%")
            except Exception as prod_error:
                print(f"Error processing VAT product: {prod_error}")
                continue
        
        print(f"Built VAT lookup with {len(vat_lookup)} entries")
        print(f"VAT lookup keys: {list(vat_lookup.keys())}")
        
        # 3. Identify relevant columns in the DataFrame
        available_columns = [col.lower() for col in df.columns]
        print(f"Available columns (lowercase): {available_columns}")

        # Try to find the right column names for product_type, country, net_price, shipping_amount, currency, order_date
        product_type_col = None
        country_col = None
        net_price_col = None
        shipping_amount_col = None
        currency_col = None
        order_date_col = None

        #For internal testing
        currencies = []
        net_prices = []

        # Find columns by name (case-insensitive)
        for col in df.columns:
            col_lower = col.lower()
            if col_lower == 'order_date':
                order_date_col = col
            elif col_lower == 'product_type':
                product_type_col = col
            elif col_lower == 'country':
                country_col = col
            elif col_lower == 'net_price':
                net_price_col = col
                net_prices = df['net_price']
            elif col_lower == 'shipping_amount':
                shipping_amount_col = col
            elif col_lower == 'currency':
                currency_col = col
                currencies = df['currency']

        # 4. Prepare lists to store calculated values for new columns
        vat_rates = []  
        vat_amounts = []
        shipping_vat_rates = []
        shipping_vat_amounts = []
        total_vat_amounts = []
        gross_total_amounts = []

        # Lists for converted prices and currencies
        converted_prices = []
        converted_shipping_prices = []
        final_currencies = []

        vat_lookup_status = []
        debug_info = []

        # 5. Fetch ECB currency rates for conversion
        ecb_rates = await get_ecb_fx_rates()

        # Track rows that need manual review
        manual_review_rows = []
        
        # 6. Process each row in the DataFrame
        for idx, row in df.iterrows():
            try:
                # Get currency and order date for this row
                currency = str(row[currency_col]).strip().upper() if currency_col else "EUR"
                order_date = str(row[order_date_col]).strip() if order_date_col else None

                # Extract product type and country
                product_type = str(row[product_type_col]).strip().lower() if product_type_col else str(row.get("product_type", "")).strip().lower()
                country = str(row[country_col]).strip().lower() if country_col else str(row.get("country", "")).strip().lower()

                # Extract net price and shipping amount
                net_price = safe_float(row[net_price_col]) if net_price_col else safe_float(row.get("price", 0))
                shipping_amount = safe_float(row[shipping_amount_col]) if shipping_amount_col else safe_float(row.get("shipping_amount", 0))
                
                # 7. Convert currency to EUR if needed
                fx_rate = None
                if currency != "EUR" and order_date:
                    order_date_str = pd.to_datetime(order_date).strftime('%Y-%m-%d')
                    fx_rate = get_fx_rate_for_date(ecb_rates, order_date_str, currency)
                    if fx_rate:
                        net_price = safe_round(net_price / fx_rate, 2)
                        shipping_amount = safe_round(shipping_amount / fx_rate, 2)
                        print(f"Row {idx}: Converted from {currency} to EUR using rate {fx_rate} for date {order_date}")
                    else:
                        print(f"Row {idx}: No FX rate found for {currency} on or before {order_date}, using original values")

                converted_prices.append(net_price)
                converted_shipping_prices.append(shipping_amount)
                final_currencies.append("EUR" if currency != "EUR" and fx_rate else currency)
                
                total_net_price = safe_round(sum(converted_prices), 2)

                print("Converted Prices", converted_prices)
                print(f"Row {idx}: product_type='{product_type}', country='{country}', price={net_price}, shipping_price={shipping_amount}")
                
                # 8. Look up VAT data for this product_type and country
                lookup_key = (product_type, country)
                vat_data = vat_lookup.get(lookup_key)

                if vat_data:
                    # Extract VAT rates and calculate VAT amounts
                    vat_rate = safe_float(vat_data.get("vat_rate", 2))
                    vat_rate = safe_round((vat_rate / 100), 2)
                    shipping_vat_rate = safe_float(vat_data.get("shipping_vat_rate", 2))
                    shipping_vat_rate = safe_round((shipping_vat_rate / 100), 2)
                    
                    vat_amount = safe_round(vat_rate * net_price, 2)
                    shipping_vat_amount = safe_round(shipping_vat_rate * shipping_amount, 2)
                    total_vat = safe_round(vat_amount + shipping_vat_amount, 2)
                    gross_total = safe_round(net_price + vat_amount + shipping_vat_amount, 2)

                    lookup_status = "Found"
                    debug_msg = f"VAT found: {vat_rate}% VAT, {shipping_vat_rate}% shipping VAT"
                    print(f"Row {idx}: {debug_msg}")

                else:
                    # No VAT data found - use defaults and log for manual review
                    manual_review_rows.append(row.to_dict())
                    vat_rate = "Not Found"
                    shipping_vat_rate = "Not Found"
                    vat_amount = "Not Found"
                    shipping_vat_amount = "Not Found"
                    gross_total = 0.0
                    total_vat = 0.0
                    lookup_status = "Not Found"
                    debug_msg = f"No VAT data found for key: {lookup_key}"
                    print(f"Row {idx}: {debug_msg}")
                
                # Store calculated values for this row
                vat_rates.append(vat_rate)
                vat_amounts.append(vat_amount)
                shipping_vat_rates.append(shipping_vat_rate)
                shipping_vat_amounts.append(shipping_vat_amount)
                total_vat_amounts.append(total_vat)
                gross_total_amounts.append(gross_total)
                vat_lookup_status.append(lookup_status)
                debug_info.append(debug_msg)
                
            except Exception as row_error:
                print(f"Error processing row {idx}: {str(row_error)}")
                manual_review_rows.append(idx)
                # Use safe defaults for this row
                vat_rates.append(0.0)
                vat_amounts.append(0.0)
                shipping_vat_rates.append(0.0)
                shipping_vat_amounts.append(0.0)
                total_vat_amounts.append(0.0)
                gross_total_amounts.append(0.0)
                vat_lookup_status.append("Error")
                debug_info.append(f"Error: {str(row_error)}")
                
        # 9. Update DataFrame with converted prices and currencies
        if net_price_col:
            df[net_price_col] = converted_prices
        if shipping_amount_col:
            df[shipping_amount_col] = converted_shipping_prices
        if currency_col:
            df[currency_col] = final_currencies

        # 10. Add new VAT-related columns to the DataFrame
        df["Previous Currency"] = currencies
        df["Previous Net Price"] = net_prices
        df["VAT Rate"] = vat_rates
        df["Product VAT"] = vat_amounts
        df["Shipping VAT Rate"] = shipping_vat_rates
        df["Shipping VAT"] = shipping_vat_amounts
        df["Total VAT"] = total_vat_amounts
        df["Final Gross Total"] = gross_total_amounts
        # df.loc[0, "Overall VAT Amount"] = safe_round(sum(total_vat_amounts), 2)
        # df.loc[0, "Overall Net Price"] = total_net_price
        # df.loc[0, "Overall Gross Total"] = safe_float(sum(gross_total_amounts), 2)

        # 11. Optionally, rename columns to user-friendly labels from header config
        all_headers = await get_all_headers()
        header_labels = {}
        for header in all_headers:
            value = header['value']
            label = header['label']
            header_labels[value] = label
        
        rename_map = {}
        for col in df.columns:
            if col in header_labels:
                rename_map[col] = header_labels[col]
                print(f"Will rename column '{col}' to '{header_labels[col]}'")
        
        if rename_map:
            df.rename(columns=rename_map, inplace=True)
        else:
            print("No columns found that match header values for renaming")
        
        # 12. Print final DataFrame for debugging
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', None)
        print("DataFrame table (full view with renamed headers):")
        print(df)

        # 13. Create a summary VAT report by country
        summary = df.groupby('Country').agg({
            'Net Price': 'sum',
            'Total VAT': 'sum'
        }).reset_index()

        summary.rename(columns={
            'Net Price': 'Net Sales',
            'Total VAT': 'VAT Amount'
        }, inplace=True)

        print("Summary VAT Report by Country:")
        print(summary)
        print("Manual Review", manual_review_rows)
        manual_df = pd.DataFrame(manual_review_rows)
        print("Manual DataFrame", manual_df)

        # 14. Return the enriched DataFrame and summary DataFrame
        if len(manual_review_rows) > 0:
            return {
                "status": "manual_review_required",
                "message": "Some rows could not be processed automatically. We'll email you the results within 24 hours.",
                "manual_review_count": len(manual_review_rows),
                "require_email": True  # Frontend will use this flag to prompt the user
            }

        # At the end of enrich_dataframe_with_vat
        return (
                df,
                summary,
                manual_df,
                {
                    "overall_vat_amount": safe_round(sum(total_vat_amounts), 2),
                    "overall_net_price": total_net_price,
                    "overall_gross_total": safe_float(sum(gross_total_amounts), 2)
                }
            )

        
    except Exception as e:
        print(f"Error in VAT enrichment: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to enrich data with VAT: {str(e)}")
    
@router.post("/validate-file")
async def validate_file(files: List[UploadFile] = File(...)):
    results = []
    for file in files:
        try:
            print(f"Processing file: {file.filename}")
            
            # Check file type
            allowed_extensions = ['.csv', '.txt', '.xls', '.xlsx']
            file_extension = '.' + file.filename.split('.')[-1].lower()
            if file_extension not in allowed_extensions:
                results.append({
                    "file_name": file.filename,
                    "success": False,
                    "message": f"Unsupported file type: {file_extension}"
                })
                continue
            
            # Extract headers and data from file
            headers, df = await extract_file_headers(file)
            
            if not headers:
                results.append({
                    "file_name": file.filename,
                    "success": False,
                    "message": "No headers found in the file"
                })
                continue
            
            # Validate file data
            validation_result = await validate_file_data(headers, df)
            print("File validation completed")
            
            # Enrich with VAT data
            # try:
            #     enriched_df = await enrich_dataframe_with_vat(df)
            #     print("VAT enrichment completed successfully")
                
            #     # Show sample of enriched data
            #     # print("Sample enriched data:")
            #     # vat_columns = ['vat_rate', 'vat_amount', 'shipping_vat_rate', 'shipping_vat_amount', 'total_vat', 'vat_lookup_status']
            #     # available_vat_cols = [col for col in vat_columns if col in enriched_df.columns]
            #     # if available_vat_cols:
            #     #     print(enriched_df[available_vat_cols].head().to_string())
                
            #     # # Convert to JSON-safe format
            #     # # Convert full enriched DataFrame to JSON-safe format
            #     # enriched_df = enriched_df.where(pd.notnull(enriched_df), None)
            #     # validation_result["enriched_data"] = dataframe_to_json_safe(enriched_df)


            #     validation_result["enrichment_success"] = True
            #     validation_result["enrichment_message"] = f"Successfully enriched {len(validation_result['enriched_data'])} rows with VAT data"

            # except Exception as enrich_error:
            #     print(f"VAT enrichment failed: {str(enrich_error)}")
            #     import traceback
            #     traceback.print_exc()
                
            #     validation_result["enrichment_success"] = False
            #     validation_result["enrichment_message"] = f"VAT enrichment failed: {str(enrich_error)}"
            
            has_issues = len(validation_result['missing_headers']) > 0 or len(validation_result['data_issues']) > 0
            
            results.append({
                "file_name": file.filename,
                "success": not has_issues,
                "has_issues": has_issues,
                "validation_result": validation_result,
                "message": "File has validation issues" if has_issues else "File validation completed successfully"
            })
            
        except Exception as e:
            print(f"Error processing file {file.filename}: {str(e)}")
            import traceback
            traceback.print_exc()
            results.append({
                "file_name": file.filename,
                "success": False,
                "message": f"Error validating file: {str(e)}"
            })

    return {"files": results}

@router.post("/download-vat-issues")
async def download_vat_issues(file: UploadFile = File(...)):
    try:
        headers, df = await extract_file_headers(file)
        validation_result = await validate_file_data(headers, df)

        for col in df.columns:
            if "order date" in col.lower():
                df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d-%m-%Y")
        # Get header labels mapping
        all_headers = await get_all_headers()
        header_labels = {}
        
        # Build mapping from header values to labels
        for header in all_headers:
            value = header['value']
            label = header['label']
            header_labels[value] = label
        
        reverse_rename_map = {}
        for key, val in header_labels.items():
            reverse_rename_map[key] = val
        
        # Create rename mapping for existing columns
        rename_map = {}
        for col in df.columns:
            # Check if this column has a corresponding label
            if col in header_labels:
                rename_map[col] = header_labels[col]
        
        # Apply the renaming
        if rename_map:
            df.rename(columns=rename_map, inplace=True)
        

        issues = validation_result.get("data_issues", [])
        missing_headers = validation_result.get("missing_headers_detailed", [])
        header_labels = validation_result.get("header_labels", {})

        # Insert placeholder columns for missing headers
        missing_labels = [mh["header_label"] for mh in missing_headers]
        for label in missing_labels:
            if label not in df.columns:
                df[label] = ""

        # Build issues sheet
        issues_rows = []
        for mh in missing_headers:
            issues_rows.append({
                "Issue Type": "Missing Header",
                "Column": mh["header_label"],
                "Description": mh["description"]
            })
        for issue in issues:
            issues_rows.append({
                "Issue Type": issue["issue_type"],
                "Column": issue["column_name"],
                "Description": issue["issue_description"],
                "Missing Count": issue.get("total_missing", ""),
                "Missing %": issue.get("percentage", "")
            })
        issues_df = pd.DataFrame(issues_rows or [{
            "Issue Type": "None",
            "Column": "All",
            "Description": "No missing headers or data issues found."
        }])

        # --- Create Excel workbook ---
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "User Data"

        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Fills
        red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")    # Header or invalid type
        orange_fill = PatternFill(start_color="FFBF00", end_color="FFF2CC", fill_type="solid")  # Missing values
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Normal header
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        # --- Write data ---
        for r in dataframe_to_rows(df, index=False, header=True):
            ws_data.append(r)

        col_name_to_index = {col: idx for idx, col in enumerate(df.columns)}

        # --- Highlight headers ---
        for col_idx, col in enumerate(df.columns, start=1):
            cell = ws_data.cell(row=1, column=col_idx)
            cell.font = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            cell.fill = red_fill if col in missing_labels else header_fill

        # --- Highlight issues ---
        for issue in issues:
            original_col = issue.get("original_column")
            if not original_col:
                continue

            # Map system name to label
            renamed_col = reverse_rename_map.get(original_col, original_col)

            if renamed_col not in col_name_to_index:
                continue

            col_idx = col_name_to_index[renamed_col] + 1  # openpyxl is 1-indexed

            if issue["issue_type"] == "MISSING_DATA":
                for row_str in issue.get("missing_rows", []):
                    try:
                        row_num = int(row_str)
                        ws_data.cell(row=row_num, column=col_idx).fill = orange_fill
                    except Exception:
                        continue

            elif issue["issue_type"] == "INVALID_TYPE":
                for row_num in issue.get("invalid_rows", []):
                    try:
                        ws_data.cell(row=row_num, column=col_idx).fill = red_fill
                    except Exception:
                        continue


        # --- Style data cells + autosize ---
        for row in ws_data.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        for col in ws_data.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            col_letter = get_column_letter(col[0].column)
            ws_data.column_dimensions[col_letter].width = max_len + 2

        # --- Add issues sheet ---
        ws_issues = wb.create_sheet("Validation Issues")
        for r in dataframe_to_rows(issues_df, index=False, header=True):
            ws_issues.append(r)

        for col in ws_issues.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.fill = header_fill
                cell.font = bold_font
                cell.border = thin_border

        for row in ws_issues.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col in ws_issues.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            col_letter = get_column_letter(col[0].column)
            ws_issues.column_dimensions[col_letter].width = max_len + 2

        # --- Output stream ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        download_name = file.filename.rsplit('.', 1)[0] + "_validation_annotated.xlsx"
        return StreamingResponse(
            output,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={download_name}"}
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Could not generate issue report: {str(e)}")
  

@router.post("/download-vat-report")
async def download_vat_report(background_tasks: BackgroundTasks, file: UploadFile = File(...), user_email: str = Form(...)):
    try:
        headers, df = await extract_file_headers(file)
        await validate_file_data(headers, df)
        print("File validation completed")

        # Call enrichment function
        result = await enrich_dataframe_with_vat(df)
        print("Enrichment result:", result)

        # Handle manual review case where enrichment directly returns a dict
        if isinstance(result, dict):
            if result.get("status") == "manual_review_required":
                print("Manual review required. Preparing email.")
                
                if user_email: 
                    # Rebuild the manual email file again
                    manual_email_stream = io.BytesIO()
                    with pd.ExcelWriter(manual_email_stream, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name="VAT Report")
                        # summary_df.to_excel(writer, index=False, sheet_name="Summary")

                    manual_email_stream.seek(0)

                    # Modify font and highlight rows with "Not Found" in any cell
                    workbook = load_workbook(manual_email_stream)
                    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                    for sheet_name in ["VAT Report", "Summary"]:
                        if sheet_name in workbook.sheetnames:
                            sheet = workbook[sheet_name]
                            
                            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                                highlight_row = any(str(cell.value).strip() == "Not Found" for cell in row)
                                for cell in row:
                                    cell.font = Font(name='Calibri', size=12, bold=False)
                                    if highlight_row:
                                        cell.fill = fill

                            # Set font for header row (row 1)
                            for cell in sheet[1]:
                                cell.font = Font(name='Calibri', size=12, bold=False)


                    # Re-save to a new stream
                    final_manual_email_stream = io.BytesIO()
                    workbook.save(final_manual_email_stream)
                    final_manual_email_stream.seek(0)

                    background_tasks.add_task(
                        send_manual_vat_email,
                        "anandpandey1765@gmail.com",  # Admin
                        user_email,                   # User for context
                        final_manual_email_stream.getvalue()
                    )
                    print("Manual review email task added to background.")
                else:
                    print("No user email provided. Skipping manual review email.")

                if "file_name" not in result:
                    result["file_name"] = file.filename

                return JSONResponse(status_code=200, content=result)


        # Otherwise, continue with file generation for download and potential email
        enriched_df, summary_df, manual_df, vat_summary = result

        # Format 'order date' column if it exists and is a datetime
        for col in enriched_df.columns:
            if "order date" in col.lower() and pd.api.types.is_datetime64_any_dtype(enriched_df[col]):
                enriched_df[col] = enriched_df[col].dt.strftime("%d-%m-%Y")

        # --- Generate the Excel file for direct download ---
        download_excel_stream = io.BytesIO()
        with pd.ExcelWriter(download_excel_stream, engine='openpyxl') as writer:
            enriched_df.to_excel(writer, index=False, sheet_name="VAT Report")
            summary_df.to_excel(writer, index=False, sheet_name="Summary")

            workbook = writer.book
            vat_report_sheet = writer.sheets["VAT Report"]

            # Leave one empty row after the data
            start_row = enriched_df.shape[0] + 3

            # Write labeled totals below the data
            vat_report_sheet.cell(row=start_row, column=5, value="Overall Net Total")
            vat_report_sheet.cell(row=start_row + 1, column=5, value=vat_summary["overall_net_price"])

            vat_report_sheet.cell(row=start_row, column=13, value="Overall VAT Amount")
            vat_report_sheet.cell(row=start_row + 1, column=13, value=vat_summary["overall_vat_amount"])

            vat_report_sheet.cell(row=start_row, column=14, value="Overall Gross Total")
            vat_report_sheet.cell(row=start_row + 1, column=14, value=vat_summary["overall_gross_total"])


        # Highlight "Not Found" rows in the downloadable report
        download_excel_stream.seek(0)
        workbook = load_workbook(download_excel_stream)

        # Apply formatting to both sheets
        for sheet_name in ["VAT Report", "Summary"]:
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # Set font size 12, no bold
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.font = Font(name='Calibri', size=12, bold=False)

        final_download_stream = io.BytesIO()
        workbook.save(final_download_stream)
        final_download_stream.seek(0)
        print("Manual Df is empty", manual_df.empty)

        download_name = file.filename.rsplit('.', 1)[0] + "_vat_report.xlsx"
        return StreamingResponse(
            final_download_stream,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={download_name}"}
        )

    except HTTPException as he:
        print(f"HTTPException caught: {he.detail}")
        raise he # Re-raise HTTPException to be handled by FastAPI
    except Exception as e:
        print(f"Unhandled exception in download_vat_report: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Could not generate download: {str(e)}")
