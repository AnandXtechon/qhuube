from io import BytesIO
import io
from typing import List, Dict, Any, Optional
import pandas as pd
from fastapi import BackgroundTasks, Form, UploadFile, HTTPException, APIRouter, File
from fastapi.responses import JSONResponse, StreamingResponse
from app.models.header_model import get_all_headers
from app.models.product_model import get_all_products
from app.core.helper import rename_columns_with_labels, safe_float, safe_round, dataframe_to_json_safe, get_user_friendly_dtype, TYPE_MAP
from app.core.currency_conversion import get_ecb_fx_rates, get_fx_rate_for_date
from app.core.send_mail import send_manual_vat_email
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import re
import uuid
from datetime import datetime, timedelta
import asyncio
from concurrent.futures import ThreadPoolExecutor
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

router = APIRouter()

# Enhanced in-memory storage for processed data
processed_data_store: Dict[str, Dict[str, Any]] = {}

# Thread pool for CPU-intensive operations
executor = ThreadPoolExecutor(max_workers=4)

class DataProcessor:
    """Handles all data processing operations with caching"""
    
    def __init__(self):
        self.vat_lookup_cache = None
        self.headers_cache = None
        self.ecb_rates_cache = None
        self.cache_timestamp = None
        
    async def get_cached_vat_lookup(self):
        """Get cached VAT lookup or fetch if expired"""
        current_time = datetime.now()
        if (self.vat_lookup_cache is None or 
            self.cache_timestamp is None or 
            current_time - self.cache_timestamp > timedelta(hours=1)):
            
            logger.info("Refreshing VAT lookup cache")
            vat_products = await get_all_products()
            self.vat_lookup_cache = {}
            
            for prod in vat_products:
                try:
                    product_type = str(prod.get('product_type', '')).strip().lower()
                    country = str(prod.get('country', '')).strip().lower()
                    if product_type and country:
                        key = (product_type, country)
                        self.vat_lookup_cache[key] = prod
                except Exception as e:
                    logger.error(f"Error processing VAT product: {e}")
                    continue
            
            self.cache_timestamp = current_time
            logger.info(f"VAT lookup cache refreshed with {len(self.vat_lookup_cache)} entries")
        
        return self.vat_lookup_cache
    
    async def get_cached_headers(self):
        """Get cached headers or fetch if expired"""
        if self.headers_cache is None:
            self.headers_cache = await get_all_headers()
        return self.headers_cache
    
    async def get_cached_ecb_rates(self):
        """Get cached ECB rates or fetch if expired"""
        if self.ecb_rates_cache is None:
            self.ecb_rates_cache = await get_ecb_fx_rates()
        return self.ecb_rates_cache

# Global data processor instance
data_processor = DataProcessor()

def cleanup_old_data():
    """Cleanup old entries (older than 2 hours)"""
    current_time = datetime.now()
    expired_keys = []
    for key, data in processed_data_store.items():
        if current_time - data['timestamp'] > timedelta(hours=2):
            expired_keys.append(key)
    for key in expired_keys:
        del processed_data_store[key]
    if expired_keys:
        logger.info(f"Cleaned up {len(expired_keys)} expired entries")

async def extract_file_headers(file: UploadFile) -> tuple[list[str], pd.DataFrame]:
    """Extract headers and DataFrame from uploaded file"""
    try:
        content = await file.read()
        file_data = BytesIO(content)
        
        if file.filename.endswith('.csv'):
            df = pd.read_csv(file_data)
        elif file.filename.endswith('.txt'):
            df = pd.read_csv(file_data, delimiter='\t')
        else:
            df = pd.read_excel(file_data)
        
        headers = [str(col).strip().lower() for col in df.columns]
        return headers, df
    except Exception as e:
        logger.error(f"Error reading file {file.filename}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Error reading file: {str(e)}"
        )

async def validate_file_data(file_headers: list[str], df: pd.DataFrame) -> dict:
    """Validate file data against expected headers and types"""
    try:
        all_headers = await data_processor.get_cached_headers()
        alias_to_value = {}
        required_headers = []
        header_labels = {}
        expected_types = {}
        
        for header in all_headers:
            value = header['value']
            required_headers.append(value)
            header_labels[value] = header['label']
            for alias in header['aliases']:
                alias_to_value[alias.strip().lower()] = value
            
            raw_type = header.get('type', 'string')
            mapped_type = TYPE_MAP.get(raw_type.lower(), 'string')
            expected_types[value] = mapped_type

        # Map columns to standard names
        rename_map = {}
        for col in df.columns:
            normalized = col.strip().lower()
            mapped_value = alias_to_value.get(normalized)
            if mapped_value:
                rename_map[col] = mapped_value

        df.rename(columns=rename_map, inplace=True)
        
        # Identify missing headers
        missing_headers_detailed = []
        for field in required_headers:
            if field not in df.columns:
                missing_headers_detailed.append({
                    'header_value': field,
                    'header_label': header_labels.get(field, field),
                    'expected_name': header_labels.get(field, field),
                    'description': f"Required column '{header_labels.get(field, field)}' is missing from the file"
                })

        # Validate data quality
        data_issues = await validate_data_quality(df, header_labels, expected_types)
        
        return {
            'missing_headers': [field for field in required_headers if field not in df.columns],
            'missing_headers_detailed': missing_headers_detailed,
            'matched_columns': {v: v for v in df.columns},
            'header_labels': header_labels,
            'data_issues': data_issues,
            'total_rows': len(df),
        }
    except Exception as e:
        logger.error(f"Validation error: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Validation error: {str(e)}"
        )

async def validate_data_quality(df: pd.DataFrame, header_labels: dict, expected_types: dict) -> list:
    """Validate data quality issues in parallel"""
    data_issues = []
    
    # Process columns in parallel for better performance
    tasks = []
    for header_value in df.columns:
        if header_value in df.columns:
            task = validate_column_data(df, header_value, header_labels, expected_types)
            tasks.append(task)
    
    if tasks:
        column_issues = await asyncio.gather(*tasks, return_exceptions=True)
        for issues in column_issues:
            if isinstance(issues, list):
                data_issues.extend(issues)
            elif isinstance(issues, Exception):
                logger.error(f"Error validating column: {issues}")
    
    return data_issues

async def validate_column_data(df: pd.DataFrame, header_value: str, header_labels: dict, expected_types: dict) -> list:
    """Validate a single column's data quality"""
    issues = []
    
    try:
        # Check for missing data
        null_mask = df[header_value].isnull()
        empty_mask = df[header_value].astype(str).str.strip().isin(['', 'nan', 'None', '(empty)', '(null)'])
        combined_mask = null_mask | empty_mask
        total_empty = int(combined_mask.sum())
        
        if total_empty > 0:
            missing_rows = df.index[combined_mask].tolist()
            missing_rows_display = [str(row + 2) for row in missing_rows[:10]]
            
            issue_description = f"Column '{header_labels.get(header_value, header_value)}' has {total_empty} missing values"
            if len(missing_rows) > 10:
                issue_description += f" (showing first 10 rows: {','.join(missing_rows_display)}...)"
            else:
                issue_description += f" in rows: {', '.join(missing_rows_display)}"
            
            issues.append({
                'header_value': header_value,
                'header_label': header_labels.get(header_value, header_value),
                'original_column': header_value,
                'issue_type': 'MISSING_DATA',
                'issue_description': issue_description,
                'column_name': header_labels.get(header_value, header_value),
                'total_missing': total_empty,
                'percentage': round((total_empty / len(df)) * 100, 2),
                'missing_rows': missing_rows_display,
                'has_more_rows': len(missing_rows) > 10
            })
        
        # Check for type validation issues
        expected_type = expected_types.get(header_value, "string")
        invalid_rows = await validate_column_types(df[header_value], expected_type)
        
        if invalid_rows:
            invalid_rows_display = invalid_rows[:10]
            issue_description = f"Column '{header_labels.get(header_value, header_value)}' has invalid {expected_type} values in rows: {', '.join(map(str, invalid_rows_display))}"
            if len(invalid_rows) > 10:
                issue_description += "..."
            
            issues.append({
                'header_value': header_value,
                'header_label': header_labels.get(header_value, header_value),
                'original_column': header_value,
                'issue_type': 'INVALID_TYPE',
                'issue_description': issue_description,
                'column_name': header_labels.get(header_value, header_value),
                'expected_type': expected_type,
                'invalid_rows': invalid_rows_display,
                'invalid_count': len(invalid_rows),
                'total_rows': len(df),
                'percentage': round((len(invalid_rows) / len(df)) * 100, 2),
                'has_more_rows': len(invalid_rows) > 10
            })
    
    except Exception as e:
        logger.error(f"Error validating column {header_value}: {str(e)}")
    
    return issues

async def validate_column_types(series: pd.Series, expected_type: str) -> list:
    """Validate column types and return invalid row numbers"""
    invalid_rows = []
    
    for idx, val in series.items():
        if pd.isnull(val) or (isinstance(val, str) and val.strip() == ''):
            continue
        
        is_valid = True
        
        try:
            if expected_type == 'integer':
                if isinstance(val, bool):
                    is_valid = False
                elif isinstance(val, str):
                    clean_val = val.strip()
                    if not clean_val.replace('-', '').replace('+', '').isdigit():
                        is_valid = False
                else:
                    float_val = float(val)
                    if float_val != int(float_val):
                        is_valid = False
            
            elif expected_type == 'float':
                try:
                    float(val)
                except (ValueError, TypeError):
                    is_valid = False
            
            elif expected_type == 'date':
                date_formats = ["%d-%m-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"]
                date_parsed = False
                
                for fmt in date_formats:
                    try:
                        pd.to_datetime(str(val), format=fmt, errors='raise')
                        date_parsed = True
                        break
                    except:
                        continue
                
                if not date_parsed:
                    try:
                        pd.to_datetime(str(val), errors='raise')
                        date_parsed = True
                    except:
                        is_valid = False
            
            elif expected_type == 'text_only':
                str_val = str(val).strip()
                if re.match(r'^-?\d+\.?\d*$', str_val):
                    is_valid = False
            
            elif expected_type == 'string':
                str_val = str(val).strip()
                # Add specific validation rules based on column context
                if len(str_val) == 0:
                    is_valid = False
        
        except Exception:
            is_valid = False
        
        if not is_valid:
            invalid_rows.append(idx + 2)  # +2 for 1-indexed + header row
    
    return invalid_rows

async def enrich_dataframe_with_vat(df: pd.DataFrame) -> tuple:
    """Enrich DataFrame with VAT calculations using cached data"""
    try:
        logger.info("Starting VAT enrichment process")
        
        # Get cached data
        vat_lookup = await data_processor.get_cached_vat_lookup()
        ecb_rates = await data_processor.get_cached_ecb_rates()
        
        # Identify relevant columns
        column_mapping = {
            'product_type': None,
            'country': None,
            'net_price': None,
            'shipping_amount': None,
            'currency': None,
            'order_date': None
        }
        
        for col in df.columns:
            col_lower = col.lower()
            if col_lower == 'order_date':
                column_mapping['order_date'] = col
            elif col_lower == 'product_type':
                column_mapping['product_type'] = col
            elif col_lower == 'country':
                column_mapping['country'] = col
            elif col_lower == 'net_price':
                column_mapping['net_price'] = col
            elif col_lower == 'shipping_amount':
                column_mapping['shipping_amount'] = col
            elif col_lower == 'currency':
                column_mapping['currency'] = col
        
        # Process rows in batches for better performance
        batch_size = 1000
        total_rows = len(df)
        
        # Initialize result lists
        vat_rates = []
        vat_amounts = []
        shipping_vat_rates = []
        shipping_vat_amounts = []
        total_vat_amounts = []
        gross_total_amounts = []
        converted_prices = []
        converted_shipping_prices = []
        final_currencies = []
        vat_lookup_status = []
        manual_review_rows = []
        
        # Store original values for reference
        original_currencies = df[column_mapping['currency']].tolist() if column_mapping['currency'] else ['EUR'] * total_rows
        original_net_prices = df[column_mapping['net_price']].tolist() if column_mapping['net_price'] else [0.0] * total_rows
        
        logger.info(f"Processing {total_rows} rows in batches of {batch_size}")
        
        for batch_start in range(0, total_rows, batch_size):
            batch_end = min(batch_start + batch_size, total_rows)
            batch_df = df.iloc[batch_start:batch_end]
            
            logger.info(f"Processing batch {batch_start}-{batch_end}")
            
            # Process batch
            batch_results = await process_vat_batch(
                batch_df, column_mapping, vat_lookup, ecb_rates, batch_start
            )
            
            # Extend result lists
            vat_rates.extend(batch_results['vat_rates'])
            vat_amounts.extend(batch_results['vat_amounts'])
            shipping_vat_rates.extend(batch_results['shipping_vat_rates'])
            shipping_vat_amounts.extend(batch_results['shipping_vat_amounts'])
            total_vat_amounts.extend(batch_results['total_vat_amounts'])
            gross_total_amounts.extend(batch_results['gross_total_amounts'])
            converted_prices.extend(batch_results['converted_prices'])
            converted_shipping_prices.extend(batch_results['converted_shipping_prices'])
            final_currencies.extend(batch_results['final_currencies'])
            vat_lookup_status.extend(batch_results['vat_lookup_status'])
            manual_review_rows.extend(batch_results['manual_review_rows'])
        
        # Update DataFrame with results
        if column_mapping['net_price']:
            df[column_mapping['net_price']] = converted_prices
        if column_mapping['shipping_amount']:
            df[column_mapping['shipping_amount']] = converted_shipping_prices
        if column_mapping['currency']:
            df[column_mapping['currency']] = final_currencies
        
        # Add new columns
        df["Previous Currency"] = original_currencies
        df["Previous Net Price"] = original_net_prices
        df["VAT Rate"] = vat_rates
        df["Product VAT"] = vat_amounts
        df["Shipping VAT Rate"] = shipping_vat_rates
        df["Shipping VAT"] = shipping_vat_amounts
        df["Total VAT"] = total_vat_amounts
        df["Final Gross Total"] = gross_total_amounts
        
        # Rename columns to user-friendly labels
        df = await rename_columns_with_labels(df)
        
        # Create summary
        summary = df.groupby('Country').agg({
            'Net Price': 'sum',
            'Total VAT': 'sum'
        }).reset_index()
        summary.rename(columns={
            'Net Price': 'Net Sales',
            'Total VAT': 'VAT Amount'
        }, inplace=True)
        
        # Handle manual review cases
        if manual_review_rows:
            manual_df = pd.DataFrame(manual_review_rows)
            manual_df = await rename_columns_with_labels(manual_df)
            
            return {
                "status": "manual_review_required",
                "message": "Some rows could not be processed automatically. We'll email you the results within 24 hours.",
                "manual_review_count": len(manual_review_rows),
                "require_email": True,
                "manual_review_rows": df.to_dict(orient="records"),
            }
        
        # Calculate totals
        vat_summary = {
            "overall_vat_amount": safe_round(sum(total_vat_amounts), 2),
            "overall_net_price": safe_round(sum(converted_prices), 2),
            "overall_gross_total": safe_round(sum(gross_total_amounts), 2)
        }
        
        logger.info("VAT enrichment completed successfully")
        return df, summary, pd.DataFrame(), vat_summary
        
    except Exception as e:
        logger.error(f"Error in VAT enrichment: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to enrich data with VAT: {str(e)}")

async def process_vat_batch(batch_df: pd.DataFrame, column_mapping: dict, vat_lookup: dict, ecb_rates: dict, batch_start: int) -> dict:
    """Process a batch of rows for VAT calculations"""
    batch_results = {
        'vat_rates': [],
        'vat_amounts': [],
        'shipping_vat_rates': [],
        'shipping_vat_amounts': [],
        'total_vat_amounts': [],
        'gross_total_amounts': [],
        'converted_prices': [],
        'converted_shipping_prices': [],
        'final_currencies': [],
        'vat_lookup_status': [],
        'manual_review_rows': []
    }
    
    for idx, row in batch_df.iterrows():
        try:
            # Extract values
            currency = str(row[column_mapping['currency']]).strip().upper() if column_mapping['currency'] else "EUR"
            order_date = str(row[column_mapping['order_date']]).strip() if column_mapping['order_date'] else None
            product_type = str(row[column_mapping['product_type']]).strip().lower() if column_mapping['product_type'] else ""
            country = str(row[column_mapping['country']]).strip().lower() if column_mapping['country'] else ""
            net_price = safe_float(row[column_mapping['net_price']]) if column_mapping['net_price'] else 0.0
            shipping_amount = safe_float(row[column_mapping['shipping_amount']]) if column_mapping['shipping_amount'] else 0.0
            
            # Currency conversion
            fx_rate = None
            if currency != "EUR" and order_date:
                order_date_str = pd.to_datetime(order_date).strftime('%Y-%m-%d')
                fx_rate = get_fx_rate_for_date(ecb_rates, order_date_str, currency)
                if fx_rate:
                    net_price = safe_round(net_price / fx_rate, 2)
                    shipping_amount = safe_round(shipping_amount / fx_rate, 2)
            
            batch_results['converted_prices'].append(net_price)
            batch_results['converted_shipping_prices'].append(shipping_amount)
            batch_results['final_currencies'].append("EUR" if currency != "EUR" and fx_rate else currency)
            
            # VAT lookup
            lookup_key = (product_type, country)
            vat_data = vat_lookup.get(lookup_key)
            
            if vat_data:
                vat_rate = safe_float(vat_data.get("vat_rate", 0)) / 100
                shipping_vat_rate = safe_float(vat_data.get("shipping_vat_rate", 0)) / 100
                
                vat_amount = safe_round(vat_rate * net_price, 2)
                shipping_vat_amount = safe_round(shipping_vat_rate * shipping_amount, 2)
                total_vat = safe_round(vat_amount + shipping_vat_amount, 2)
                gross_total = safe_round(net_price + shipping_amount + total_vat, 2)
                
                batch_results['vat_rates'].append(vat_rate)
                batch_results['vat_amounts'].append(vat_amount)
                batch_results['shipping_vat_rates'].append(shipping_vat_rate)
                batch_results['shipping_vat_amounts'].append(shipping_vat_amount)
                batch_results['total_vat_amounts'].append(total_vat)
                batch_results['gross_total_amounts'].append(gross_total)
                batch_results['vat_lookup_status'].append("Found")
            else:
                # Manual review required
                batch_results['manual_review_rows'].append(row.to_dict())
                batch_results['vat_rates'].append("Not Found")
                batch_results['vat_amounts'].append("Not Found")
                batch_results['shipping_vat_rates'].append("Not Found")
                batch_results['shipping_vat_amounts'].append("Not Found")
                batch_results['total_vat_amounts'].append(0.0)
                batch_results['gross_total_amounts'].append(0.0)
                batch_results['vat_lookup_status'].append("Not Found")
        
        except Exception as e:
            logger.error(f"Error processing row {idx}: {str(e)}")
            # Use safe defaults
            batch_results['vat_rates'].append(0.0)
            batch_results['vat_amounts'].append(0.0)
            batch_results['shipping_vat_rates'].append(0.0)
            batch_results['shipping_vat_amounts'].append(0.0)
            batch_results['total_vat_amounts'].append(0.0)
            batch_results['gross_total_amounts'].append(0.0)
            batch_results['converted_prices'].append(0.0)
            batch_results['converted_shipping_prices'].append(0.0)
            batch_results['final_currencies'].append("EUR")
            batch_results['vat_lookup_status'].append("Error")
    
    return batch_results

@router.post("/validate-file")
async def validate_file(files: List[UploadFile] = File(...)):
    """Validate uploaded files and optionally pre-process them"""
    cleanup_old_data()
    results = []
    
    for file in files:
        try:
            logger.info(f"Processing file: {file.filename}")
            
            # Check file type
            allowed_extensions = ['.csv', '.txt', '.xls', '.xlsx']
            file_extension = '.' + file.filename.split('.')[-1].lower()
            if file_extension not in allowed_extensions:
                results.append({
                    "file_name": file.filename,
                    "success": False,
                    "message": f"Unsupported file type: {file_extension}"
                })
                continue
            
            # Extract headers and data
            headers, df = await extract_file_headers(file)
            
            if not headers:
                results.append({
                    "file_name": file.filename,
                    "success": False,
                    "message": "No headers found in the file"
                })
                continue
            
            # Validate file data
            validation_result = await validate_file_data(headers, df)
            has_issues = len(validation_result['missing_headers']) > 0 or len(validation_result['data_issues']) > 0
            
            # Generate session ID
            session_id = str(uuid.uuid4())
            
            # Store data with enhanced caching
            processed_data_store[session_id] = {
                'timestamp': datetime.now(),
                'file_name': file.filename,
                'original_df': df.copy(),
                'validation_result': validation_result,
                'headers': headers,
                'has_issues': has_issues,
                'processing_status': 'validated',
                'enriched_data': None  # Will be populated when processing is requested
            }
            
            results.append({
                "file_name": file.filename,
                "session_id": session_id,
                "success": not has_issues,
                "has_issues": has_issues,
                "validation_result": validation_result,
                "message": "File has validation issues" if has_issues else "File validation completed successfully"
            })
            
        except Exception as e:
            logger.error(f"Error processing file {file.filename}: {str(e)}")
            results.append({
                "file_name": file.filename,
                "success": False,
                "message": f"Error validating file: {str(e)}"
            })
    
    return {"files": results}

@router.post("/process-vat/{session_id}")
async def process_vat_data(session_id: str):
    """Pre-process VAT data for faster downloads"""
    try:
        if session_id not in processed_data_store:
            raise HTTPException(status_code=404, detail="Session not found or expired")
        
        stored_data = processed_data_store[session_id]
        
        # Check if already processed
        if stored_data.get('processing_status') == 'enriched':
            return JSONResponse(content={
                "status": "already_processed",
                "message": "Data has already been processed"
            })
        
        # Start processing
        stored_data['processing_status'] = 'processing'
        df = stored_data['original_df'].copy()
        
        logger.info(f"Starting VAT processing for session {session_id}")
        
        # Process VAT enrichment
        result = await enrich_dataframe_with_vat(df)
        
        # Handle manual review scenario
        if isinstance(result, dict) and result.get("status") == "manual_review_required":
            stored_data['enriched_data'] = result
            stored_data['processing_status'] = 'manual_review_required'
            return JSONResponse(content=result)
        
        # Store enriched data
        enriched_df, summary_df, manual_df, vat_summary = result
        
        stored_data['enriched_data'] = {
            'enriched_df': enriched_df,
            'summary_df': summary_df,
            'manual_df': manual_df,
            'vat_summary': vat_summary,
            'status': 'completed'
        }
        stored_data['processing_status'] = 'enriched'
        
        logger.info(f"VAT processing completed for session {session_id}")
        
        return JSONResponse(content={
            "status": "completed",
            "message": "VAT processing completed successfully",
            "summary": vat_summary
        })
        
    except Exception as e:
        logger.error(f"Error processing VAT data: {str(e)}")
        if session_id in processed_data_store:
            processed_data_store[session_id]['processing_status'] = 'error'
        raise HTTPException(status_code=500, detail=f"Error processing VAT data: {str(e)}")

@router.get("/download-vat-issues/{session_id}")
async def download_vat_issues(session_id: str):
    """Download validation issues report"""
    try:
        if session_id not in processed_data_store:
            raise HTTPException(status_code=404, detail="Session not found or expired")
        
        stored_data = processed_data_store[session_id]
        df = stored_data['original_df'].copy()
        validation_result = stored_data['validation_result']
        file_name = stored_data['file_name']
        
        # Format dates
        for col in df.columns:
            if "order date" in col.lower():
                df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d-%m-%Y")
        
        # Get header labels
        all_headers = await data_processor.get_cached_headers()
        header_labels = {header['value']: header['label'] for header in all_headers}
        
        # Rename columns
        rename_map = {col: header_labels.get(col, col) for col in df.columns if col in header_labels}
        if rename_map:
            df.rename(columns=rename_map, inplace=True)
        
        # Generate Excel report
        excel_stream = await generate_issues_excel_report(df, validation_result, header_labels)
        
        download_name = file_name.rsplit('.', 1)[0] + "_validation_annotated.xlsx"
        return StreamingResponse(
            excel_stream,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={download_name}"}
        )
        
    except Exception as e:
        logger.error(f"Error generating issues report: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Could not generate issue report: {str(e)}")

async def generate_issues_excel_report(df: pd.DataFrame, validation_result: dict, header_labels: dict) -> BytesIO:
    """Generate Excel report with validation issues highlighted"""
    issues = validation_result.get("data_issues", [])
    missing_headers = validation_result.get("missing_headers_detailed", [])
    
    # Insert placeholder columns for missing headers
    missing_labels = [mh["header_label"] for mh in missing_headers]
    for label in missing_labels:
        if label not in df.columns:
            df[label] = ""
    
    # Build issues sheet data
    issues_rows = []
    for mh in missing_headers:
        issues_rows.append({
            "Issue Type": "Missing Header",
            "Column": mh["header_label"],
            "Description": mh["description"]
        })
    
    for issue in issues:
        issues_rows.append({
            "Issue Type": issue["issue_type"],
            "Column": issue["column_name"],
            "Description": issue["issue_description"],
            "Missing Count": issue.get("total_missing", ""),
            "Missing %": issue.get("percentage", "")
        })
    
    if not issues_rows:
        issues_rows = [{
            "Issue Type": "None",
            "Column": "All",
            "Description": "No missing headers or data issues found."
        }]
    
    issues_df = pd.DataFrame(issues_rows)
    
    # Create Excel workbook
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "User Data"
    
    # Styling
    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    orange_fill = PatternFill(start_color="FFBF00", end_color="FFF2CC", fill_type="solid")
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    bold_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    
    # Write data
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    
    col_name_to_index = {col: idx for idx, col in enumerate(df.columns)}
    
    # Style headers
    for col_idx, col in enumerate(df.columns, start=1):
        cell = ws_data.cell(row=1, column=col_idx)
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        cell.fill = red_fill if col in missing_labels else header_fill
    
    # Highlight issues
    reverse_rename_map = {v: k for k, v in header_labels.items()}
    
    for issue in issues:
        original_col = issue.get("original_column")
        if not original_col:
            continue
        
        renamed_col = reverse_rename_map.get(original_col, original_col)
        if renamed_col not in col_name_to_index:
            continue
        
        col_idx = col_name_to_index[renamed_col] + 1
        
        if issue["issue_type"] == "MISSING_DATA":
            for row_str in issue.get("missing_rows", []):
                try:
                    row_num = int(row_str)
                    ws_data.cell(row=row_num, column=col_idx).fill = orange_fill
                except:
                    continue
        elif issue["issue_type"] == "INVALID_TYPE":
            for row_num in issue.get("invalid_rows", []):
                try:
                    ws_data.cell(row=row_num, column=col_idx).fill = red_fill
                except:
                    continue
    
    # Style data cells and auto-size columns
    for row in ws_data.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    for col in ws_data.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = max_len + 2
    
    # Add issues sheet
    ws_issues = wb.create_sheet("Validation Issues")
    for r in dataframe_to_rows(issues_df, index=False, header=True):
        ws_issues.append(r)
    
    # Style issues sheet
    for col in ws_issues.iter_cols(min_row=1, max_row=1):
        for cell in col:
            cell.fill = header_fill
            cell.font = bold_font
            cell.border = thin_border
    
    for row in ws_issues.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    for col in ws_issues.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        col_letter = get_column_letter(col[0].column)
        ws_issues.column_dimensions[col_letter].width = max_len + 2
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

@router.post("/download-vat-report/{session_id}")
async def download_vat_report(session_id: str, background_tasks: BackgroundTasks, user_email: str = Form(...)):
    """Download VAT report using pre-processed data"""
    try:
        if session_id not in processed_data_store:
            raise HTTPException(status_code=404, detail="Session not found or expired")
        
        stored_data = processed_data_store[session_id]
        file_name = stored_data['file_name']
        
        # Check if data needs processing
        if stored_data.get('processing_status') != 'enriched':
            # Trigger processing if not done
            if stored_data.get('processing_status') != 'processing':
                await process_vat_data(session_id)
            
            # If still processing, return status
            if stored_data.get('processing_status') == 'processing':
                return JSONResponse(content={
                    "status": "processing",
                    "message": "Data is still being processed. Please try again in a moment."
                })
        
        enriched_data = stored_data.get('enriched_data')
        if not enriched_data:
            raise HTTPException(status_code=500, detail="No processed data available")
        
        # Handle manual review scenario
        if enriched_data.get("status") == "manual_review_required":
            logger.info("Manual review required. Preparing email.")
            manual_review_rows = enriched_data.get("manual_review_rows", [])
            
            # Clean timestamps for JSON serialization
            for row in manual_review_rows:
                for key, value in row.items():
                    if isinstance(value, pd.Timestamp):
                        row[key] = value.strftime("%Y-%m-%d")
            
            if user_email:
                # Generate Excel for email
                df = stored_data['original_df'].copy()
                email_stream = await generate_manual_review_excel(df)
                
                background_tasks.add_task(
                    send_manual_vat_email,
                    "anandpandey1765@gmail.com",
                    user_email,
                    email_stream.getvalue(),
                    manual_review_rows
                )
                logger.info("Manual review email task added to background.")
            
            enriched_data["file_name"] = file_name
            return JSONResponse(status_code=200, content=enriched_data)
        
        # Generate downloadable Excel file
        enriched_df = enriched_data['enriched_df']
        summary_df = enriched_data['summary_df']
        vat_summary = enriched_data['vat_summary']
        
        # Format dates
        for col in enriched_df.columns:
            if "order date" in col.lower() and pd.api.types.is_datetime64_any_dtype(enriched_df[col]):
                enriched_df[col] = enriched_df[col].dt.strftime("%d-%m-%Y")
        
        # Generate Excel report
        excel_stream = await generate_vat_excel_report(enriched_df, summary_df, vat_summary)
        
        download_name = file_name.rsplit('.', 1)[0] + "_vat_report.xlsx"
        
        # Clean up session data
        if session_id in processed_data_store:
            del processed_data_store[session_id]
        
        return StreamingResponse(
            excel_stream,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={"Content-Disposition": f"attachment; filename={download_name}"}
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error in download_vat_report: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Could not generate download: {str(e)}")

async def generate_manual_review_excel(df: pd.DataFrame) -> BytesIO:
    """Generate Excel file for manual review email"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="VAT Report")
    
    output.seek(0)
    workbook = load_workbook(output)
    
    # Apply highlighting for "Not Found" values
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            highlight_row = any(str(cell.value).strip() == "Not Found" for cell in row)
            for cell in row:
                cell.font = Font(name='Calibri', size=12, bold=False)
                if highlight_row:
                    cell.fill = fill
        
        # Style headers
        for cell in sheet[1]:
            cell.font = Font(name='Calibri', size=12, bold=True)
    
    final_output = BytesIO()
    workbook.save(final_output)
    final_output.seek(0)
    
    return final_output

async def generate_vat_excel_report(enriched_df: pd.DataFrame, summary_df: pd.DataFrame, vat_summary: dict) -> BytesIO:
    """Generate final VAT Excel report"""
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        enriched_df.to_excel(writer, index=False, sheet_name="VAT Report")
        summary_df.to_excel(writer, index=False, sheet_name="Summary")
        
        workbook = writer.book
        vat_report_sheet = writer.sheets["VAT Report"]
        
        # Add summary totals
        start_row = enriched_df.shape[0] + 3
        vat_report_sheet.cell(row=start_row, column=5, value="Overall Net Total")
        vat_report_sheet.cell(row=start_row + 1, column=5, value=vat_summary["overall_net_price"])
        vat_report_sheet.cell(row=start_row, column=13, value="Overall VAT Amount")
        vat_report_sheet.cell(row=start_row + 1, column=13, value=vat_summary["overall_vat_amount"])
        vat_report_sheet.cell(row=start_row, column=14, value="Overall Gross Total")
        vat_report_sheet.cell(row=start_row + 1, column=14, value=vat_summary["overall_gross_total"])
    
    output.seek(0)
    workbook = load_workbook(output)
    
    # Apply consistent styling
    for sheet_name in ["VAT Report", "Summary"]:
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = Font(name='Calibri', size=12, bold=False)
    
    final_output = BytesIO()
    workbook.save(final_output)
    final_output.seek(0)
    
    return final_output

@router.get("/processing-status/{session_id}")
async def get_processing_status(session_id: str):
    """Get the current processing status"""
    if session_id not in processed_data_store:
        raise HTTPException(status_code=404, detail="Session not found or expired")
    
    stored_data = processed_data_store[session_id]
    status = stored_data.get('processing_status', 'unknown')
    
    return {
        "session_id": session_id,
        "status": status,
        "timestamp": stored_data.get('timestamp'),
        "file_name": stored_data.get('file_name')
    }
